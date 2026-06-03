import base64
import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

products = [
  {"category": "Цемент и сыпучие", "name": "Цемент ПЦ М500 Д0 50 кг", "unit": "мешок", "price": 390, "old_price": None, "article": "CEM-500-50"},
  {"category": "Цемент и сыпучие", "name": "Цемент ПЦ М400 Д20 50 кг", "unit": "мешок", "price": 350, "old_price": None, "article": "CEM-400-50"},
  {"category": "Цемент и сыпучие", "name": "Песок речной мытый", "unit": "т", "price": 1200, "old_price": None, "article": "SAND-RCH"},
  {"category": "Цемент и сыпучие", "name": "Щебень фракция 20-40 мм", "unit": "т", "price": 1450, "old_price": None, "article": "GRAVEL-2040"},
  {"category": "Цемент и сыпучие", "name": "Отсев гранитный 0-5 мм", "unit": "т", "price": 980, "old_price": None, "article": "GRANIT-05"},
  {"category": "Цемент и сыпучие", "name": "Керамзит фракция 10-20 мм", "unit": "м³", "price": 2800, "old_price": None, "article": "KERAM-1020"},
  {"category": "Бетон и ЖБИ", "name": "Бетон товарный В15 (М200)", "unit": "м³", "price": 5200, "old_price": None, "article": "BETON-B15"},
  {"category": "Бетон и ЖБИ", "name": "Бетон товарный В25 (М350)", "unit": "м³", "price": 6100, "old_price": None, "article": "BETON-B25"},
  {"category": "Бетон и ЖБИ", "name": "Кольцо колодезное КС 10-9", "unit": "шт", "price": 2400, "old_price": None, "article": "KS-10-9"},
  {"category": "Бетон и ЖБИ", "name": "Плита перекрытия ПК 60-15", "unit": "шт", "price": 18500, "old_price": None, "article": "PK-60-15"},
  {"category": "Металл", "name": "Арматура А500С d12 мм, 11,7 м", "unit": "прут", "price": 620, "old_price": None, "article": "ARM-A500-12"},
  {"category": "Металл", "name": "Арматура А500С d8 мм, 11,7 м", "unit": "прут", "price": 280, "old_price": None, "article": "ARM-A500-8"},
  {"category": "Металл", "name": "Профильная труба 40×40×2 мм, 6 м", "unit": "шт", "price": 890, "old_price": None, "article": "PROF-40x40"},
  {"category": "Металл", "name": "Уголок 50×50×5 мм, 6 м", "unit": "шт", "price": 1150, "old_price": None, "article": "ANGLE-5050"},
  {"category": "Металл", "name": "Лист стальной г/к 2 мм, 1250×2500", "unit": "лист", "price": 4200, "old_price": None, "article": "SHEET-2MM"},
  {"category": "Лесоматериалы", "name": "Доска обрезная 25×150×6000 мм", "unit": "м³", "price": 22000, "old_price": None, "article": "BOARD-25150"},
  {"category": "Лесоматериалы", "name": "Брус строганый 100×100×6000 мм", "unit": "м³", "price": 28000, "old_price": None, "article": "BEAM-100100"},
  {"category": "Лесоматериалы", "name": "Рейка монтажная 25×50×3000 мм", "unit": "шт", "price": 95, "old_price": None, "article": "RAIL-2550"},
  {"category": "Плитные материалы", "name": "Гипсокартон Knauf ГКЛ 12,5 мм", "unit": "лист", "price": 420, "old_price": 490, "article": "GKL-125"},
  {"category": "Плитные материалы", "name": "Гипсокартон влагостойкий ГКЛВ 12,5 мм", "unit": "лист", "price": 520, "old_price": None, "article": "GKLV-125"},
  {"category": "Плитные материалы", "name": "OSB-3 плита 9 мм, 2500×1250", "unit": "лист", "price": 890, "old_price": None, "article": "OSB3-9MM"},
  {"category": "Плитные материалы", "name": "OSB-3 плита 12 мм, 2500×1250", "unit": "лист", "price": 1100, "old_price": None, "article": "OSB3-12MM"},
  {"category": "Плитные материалы", "name": "Фанера ФСФ 12 мм, 1525×1525", "unit": "лист", "price": 1350, "old_price": None, "article": "FSF-12MM"},
  {"category": "Утепление", "name": "Пенопласт ПСБ-С-25 50 мм, 1000×1000", "unit": "м²", "price": 145, "old_price": None, "article": "PSB-S25-50"},
  {"category": "Утепление", "name": "Пенопласт ПСБ-С-25 100 мм, 1000×1000", "unit": "м²", "price": 270, "old_price": None, "article": "PSB-S25-100"},
  {"category": "Утепление", "name": "Минвата ROCKWOOL Лайт Баттс 50 мм", "unit": "уп", "price": 1850, "old_price": None, "article": "RW-LB-50"},
  {"category": "Утепление", "name": "Минвата ISOVER Классик 100 мм", "unit": "уп", "price": 2100, "old_price": None, "article": "ISV-CLS-100"},
  {"category": "Гидроизоляция", "name": "Рубероид РКП-350 15 м²", "unit": "рул", "price": 380, "old_price": None, "article": "RUB-350"},
  {"category": "Гидроизоляция", "name": "Пленка п/э гидроизоляционная 200 мкм", "unit": "м²", "price": 18, "old_price": None, "article": "FILM-200"},
  {"category": "Гидроизоляция", "name": "Мастика битумная Технониколь 18 кг", "unit": "ведро", "price": 1650, "old_price": 1900, "article": "TN-MAST-18"},
  {"category": "Сухие смеси", "name": "Штукатурка гипсовая Knauf Ротбанд 30 кг", "unit": "мешок", "price": 680, "old_price": 750, "article": "KN-ROT-30"},
  {"category": "Сухие смеси", "name": "Плиточный клей Ceresit CM11 25 кг", "unit": "мешок", "price": 420, "old_price": None, "article": "CER-CM11-25"},
  {"category": "Сухие смеси", "name": "Пескобетон М300 40 кг", "unit": "мешок", "price": 195, "old_price": None, "article": "PB-M300-40"},
  {"category": "Сухие смеси", "name": "Наливной пол Bergauf Basis 25 кг", "unit": "мешок", "price": 560, "old_price": None, "article": "BG-BASIS-25"},
  {"category": "Сухие смеси", "name": "Затирка Ceresit CE33 2 кг беж", "unit": "уп", "price": 290, "old_price": None, "article": "CER-CE33-2"},
  {"category": "ЛКМ", "name": "Краска фасадная Тиккурила Бетонит 10 л", "unit": "ведро", "price": 3200, "old_price": 3600, "article": "TIK-BET-10"},
  {"category": "ЛКМ", "name": "Грунтовка глубокого проникновения 10 л", "unit": "канистра", "price": 480, "old_price": None, "article": "GRUNT-10"},
  {"category": "ЛКМ", "name": "Эмаль алкидная ПФ-115 белая 3 кг", "unit": "банка", "price": 390, "old_price": None, "article": "PF115-3KG"},
  {"category": "Тротуарная плитка", "name": "Плитка тротуарная Брусчатка 200×100×60", "unit": "м²", "price": 850, "old_price": None, "article": "PAV-BRUS-60"},
  {"category": "Тротуарная плитка", "name": "Плитка тротуарная Катушка 60 мм", "unit": "м²", "price": 920, "old_price": None, "article": "PAV-KAT-60"},
  {"category": "Тротуарная плитка", "name": "Бордюр дорожный 500×200×80", "unit": "шт", "price": 185, "old_price": None, "article": "BORD-500"},
  {"category": "Заборы и ограждения", "name": "Сетка рабица 50×50 оцинк. h=1,5 м, 10 п.м.", "unit": "рул", "price": 1650, "old_price": None, "article": "MESH-5050"},
  {"category": "Заборы и ограждения", "name": "Профнастил С8 h=2 м, шир. 1,2 м RAL 6005", "unit": "лист", "price": 680, "old_price": 750, "article": "PROF-C8-GRN"},
  {"category": "Заборы и ограждения", "name": "Столб для забора 60×60×2 мм, h=2,5 м", "unit": "шт", "price": 490, "old_price": None, "article": "POST-60-25"},
]

categories_order = [
  "Цемент и сыпучие", "Бетон и ЖБИ", "Металл", "Лесоматериалы",
  "Плитные материалы", "Утепление", "Гидроизоляция", "Сухие смеси",
  "ЛКМ", "Тротуарная плитка", "Заборы и ограждения",
]

def handler(event: dict, context) -> dict:
    """Генерирует Excel прайс-лист с товарами по категориям."""
    if event.get('httpMethod') == 'OPTIONS':
        return {
            'statusCode': 200,
            'headers': {
                'Access-Control-Allow-Origin': '*',
                'Access-Control-Allow-Methods': 'GET, OPTIONS',
                'Access-Control-Allow-Headers': 'Content-Type',
                'Access-Control-Max-Age': '86400',
            },
            'body': ''
        }

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Прайс-лист"

    header_fill = PatternFill(start_color="1A1A1A", end_color="1A1A1A", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11, name="Arial")
    cat_fill = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")
    cat_font = Font(bold=True, size=10, name="Arial")
    cell_font = Font(size=10, name="Arial")
    price_font = Font(size=10, bold=True, name="Arial")
    old_price_font = Font(size=9, name="Arial", color="999999", strike=True)
    center = Alignment(horizontal="center", vertical="center")
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    thin = Side(style="thin", color="E0E0E0")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.merge_cells("A1:F1")
    ws["A1"].value = "ПРАЙС-ЛИСТ — СтройМаркет"
    ws["A1"].font = Font(bold=True, size=14, name="Arial")
    ws["A1"].alignment = center
    ws["A1"].fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
    ws.row_dimensions[1].height = 32

    ws.merge_cells("A2:F2")
    ws["A2"].value = "Цены указаны с НДС. По вопросам опта и доставки — звоните менеджеру."
    ws["A2"].font = Font(size=9, name="Arial", color="777777", italic=True)
    ws["A2"].alignment = center
    ws.row_dimensions[2].height = 18

    headers = ["№", "Артикул", "Наименование", "Ед. изм.", "Цена продажи, руб.", "Цена до скидки, руб."]
    col_widths = [5, 18, 48, 10, 20, 20]
    for col, (h, w) in enumerate(zip(headers, col_widths), start=1):
        cell = ws.cell(row=3, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = border
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[3].height = 22

    row = 4
    num = 1
    for cat in categories_order:
        cat_products = [p for p in products if p["category"] == cat]
        if not cat_products:
            continue
        ws.merge_cells(f"A{row}:F{row}")
        cat_cell = ws.cell(row=row, column=1, value=f"  {cat.upper()}")
        cat_cell.font = cat_font
        cat_cell.fill = cat_fill
        cat_cell.alignment = left
        cat_cell.border = border
        ws.row_dimensions[row].height = 20
        row += 1
        for p in cat_products:
            ws.cell(row=row, column=1, value=num).font = cell_font
            ws.cell(row=row, column=1).alignment = center
            ws.cell(row=row, column=1).border = border
            ws.cell(row=row, column=2, value=p["article"]).font = cell_font
            ws.cell(row=row, column=2).alignment = center
            ws.cell(row=row, column=2).border = border
            ws.cell(row=row, column=3, value=p["name"]).font = cell_font
            ws.cell(row=row, column=3).alignment = left
            ws.cell(row=row, column=3).border = border
            ws.cell(row=row, column=4, value=p["unit"]).font = cell_font
            ws.cell(row=row, column=4).alignment = center
            ws.cell(row=row, column=4).border = border
            pc = ws.cell(row=row, column=5, value=p["price"])
            pc.font = price_font
            pc.alignment = center
            pc.border = border
            pc.number_format = '#,##0'
            old = p["old_price"]
            oc = ws.cell(row=row, column=6, value=old if old else "—")
            oc.font = old_price_font if old else cell_font
            oc.alignment = center
            oc.border = border
            if old:
                oc.number_format = '#,##0'
            ws.row_dimensions[row].height = 20
            num += 1
            row += 1

    ws.merge_cells(f"A{row}:D{row}")
    ws.cell(row=row, column=1, value=f"Итого позиций: {len(products)}").font = Font(bold=True, size=10, name="Arial")
    ws.cell(row=row, column=1).alignment = left
    ws.row_dimensions[row].height = 20

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    excel_b64 = base64.b64encode(buf.read()).decode("utf-8")

    return {
        'statusCode': 200,
        'headers': {
            'Access-Control-Allow-Origin': '*',
            'Content-Type': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            'Content-Disposition': 'attachment; filename="price-list.xlsx"',
        },
        'body': excel_b64,
        'isBase64Encoded': True,
    }
